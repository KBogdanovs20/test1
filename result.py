from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0


#write your code here
max_row = ws.max_row
max_salary = 3000
for row in range(2,max_row+1):
    hours = ws['B'+str(row)].value
    rate = ws['C'+str(row)].value
    if (type(hours)!=str and type(rate)!=str):
        salary=float(hours)*float(rate)
        ws['D'+str(row)].value=salary
        if salary > max_salary:
            total += 1
print(total)
wb.close()
